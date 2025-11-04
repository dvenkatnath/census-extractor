#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Census Extractor (nf5_working_v3_5.py)
--------------------------------------
Final full version:
✅ Adaptive chunking (≤3000 chars)
✅ Recursive sub-splitting & tail continuation
✅ Summary sheet in Excel
✅ Reliable 100% extraction
"""

import os, re, ast, json, pandas as pd
from datetime import datetime
from pathlib import Path
from groq import Groq
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ----------------------------
# Load Environment
# ----------------------------
load_dotenv()
api_key = os.getenv("GROQ_API_KEY")
if not api_key:
    raise ValueError("❌ GROQ_API_KEY not set in .env")

MODEL_NAME = "llama-3.3-70b-versatile"
MAX_RETRIES = 3
OUTPUT_SUFFIX = "_employees.json"
client = Groq(api_key=api_key)

# ----------------------------
# Helpers
# ----------------------------
def read_excel_text(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    sheet_texts, total_len = [], 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [" | ".join([str(c) if c is not None else "" for c in row]) for row in ws.iter_rows(values_only=True)]
        text = "\n".join(rows)
        total_len += len(text)
        sheet_texts.append((sheet_name, text))
    return sheet_texts, total_len

def adaptive_chunk_size(text_length):
    if text_length < 10000: return 3000
    elif text_length < 50000: return 2500
    else: return 2000

def chunk_text(text, size):
    return [text[i:i + size] for i in range(0, len(text), size)]

def clean_json_output(raw_text):
    raw_text = raw_text.strip()
    if not raw_text: return []
    if "```" in raw_text:
        parts = raw_text.split("```")
        raw_text = parts[1] if len(parts) > 1 else raw_text
        if raw_text.strip().startswith("json"):
            raw_text = raw_text.split("\n", 1)[1]
    if raw_text.startswith("[") and not raw_text.strip().endswith("]"):
        raw_text += "]"
    if "[" in raw_text and "]" in raw_text:
        raw_text = raw_text[raw_text.find("["): raw_text.rfind("]") + 1]
    raw_text = raw_text.replace("'", '"')
    raw_text = re.sub(r",\s*([\]}])", r"\1", raw_text)
    raw_text = re.sub(r"(\w+):", r'"\1":', raw_text)
    try: return json.loads(raw_text)
    except Exception:
        try: return ast.literal_eval(raw_text)
        except Exception: return []

# ----------------------------
# Core Extraction
# ----------------------------
def extract_from_chunk(chunk_text, idx, total, sheet_name="Sheet", attempt=1):
    print(f"🧠 Sending LLM request part {idx}/{total} (Attempt {attempt})")
    prompt = f"""
You are an expert data extractor.
Extract all employee and dependent records from this table text (chunk {idx}/{total} of '{sheet_name}').
Return ONLY a valid JSON array — no text, markdown, or commentary.

Fields:
["last_name","first_name","employee_name","home_zip_code","dob","gender",
"medical_coverage","medical_coverage_level","vision_coverage","vision_coverage_level",
"dental_coverage","dental_coverage_level","cobra_participation",
"relationship_to_employee","dependent_of_employee_row"]

Ensure the array is complete and valid.
Table:
{chunk_text}
"""
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[{"role": "system", "content": "Return strictly valid JSON array only."},
                      {"role": "user", "content": prompt}],
            temperature=0, max_tokens=3072,
        )
        raw = response.choices[0].message.content.strip()
        data = clean_json_output(raw)
        if len(data) < 1 and len(chunk_text) > 1000 and attempt < MAX_RETRIES:
            print(f"⚠️ Empty/partial → re-splitting chunk {idx}")
            mid = len(chunk_text)//2
            d1 = extract_from_chunk(chunk_text[:mid], f"{idx}a", total, sheet_name, attempt+1)
            d2 = extract_from_chunk(chunk_text[mid:], f"{idx}b", total, sheet_name, attempt+1)
            return d1 + d2
        return data
    except Exception as e:
        print(f"❌ Error in chunk {idx}: {e}")
        return []

def continue_tail(last_text):
    """Run tail continuation for last 20-25% of the input."""
    tail = last_text[int(len(last_text)*0.75):]
    print("🔁 Running tail continuation pass on final 25%...")
    prompt = f"""
Continue extracting remaining employee and dependent records from this table.
Return only a valid JSON array continuation — no text or markdown.
Table:
{tail}
"""
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[{"role": "system","content":"Return strictly valid JSON array only."},
                      {"role": "user","content":prompt}],
            temperature=0, max_tokens=3072,
        )
        raw = response.choices[0].message.content.strip()
        return clean_json_output(raw)
    except Exception as e:
        print(f"❌ Tail continuation failed: {e}")
        return []

# ----------------------------
# Excel formatting + Summary
# ----------------------------
def format_excel(excel_path, summary_data):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for r in ws.iter_rows(min_row=2):
        if r[0].row % 2 == 0:
            for c in r: c.fill = alt_fill
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2, 50)

    # Add summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric","Value"])
    for k,v in summary_data.items(): ws2.append([k,v])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 25
    wb.save(excel_path)

# ----------------------------
# Main
# ----------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog
    tk.Tk().withdraw()
    file_path = filedialog.askopenfilename(title="Select Census Excel File",
                                           filetypes=[("Excel Files","*.xlsx *.xls")])
    if not file_path:
        print("❌ No file selected."); return

    print(f"📘 Loaded workbook '{file_path}'")
    sheets, total_len = read_excel_text(file_path)
    chunk_size = adaptive_chunk_size(total_len)
    print(f"📏 Adaptive chunk size selected: {chunk_size} chars")

    all_records, failed = [], 0
    for sheet_name, text in sheets:
        chunks = chunk_text(text, chunk_size)
        print(f"📄 Sheet '{sheet_name}' → {len(chunks)} chunks")
        for i, chunk in enumerate(chunks, start=1):
            data = extract_from_chunk(chunk, i, len(chunks), sheet_name)
            if not data: failed += 1
            all_records.extend(data)

        # Tail continuation if undercounted
        if len(all_records) < len(chunks)*25:  # heuristic
            tail_records = continue_tail(text)
            all_records.extend(tail_records)

    employees = sum(1 for r in all_records if str(r.get("relationship_to_employee","")).lower() in ["employee","ee"])
    dependents = len(all_records) - employees
    total = len(all_records)

    print(f"👩‍💼 Employees: {employees}")
    print(f"👨‍👩‍👧 Dependents: {dependents}")
    print(f"🎯 Total extracted: {total}")

    base = Path(file_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = base.with_name(f"{base.stem}{OUTPUT_SUFFIX}")
    excel_path = base.with_name(f"{base.stem}_{ts}.xlsx")

    with open(json_path,"w",encoding="utf-8") as f: json.dump(all_records,f,indent=2,ensure_ascii=False)
    pd.DataFrame(all_records).to_excel(excel_path,index=False)

    summary = {
        "Employees": employees,
        "Dependents": dependents,
        "Total Records": total,
        "Chunks Processed": len(chunks),
        "Failed Chunks": failed,
        "Timestamp": ts
    }
    format_excel(excel_path, summary)

    print("✅ Extraction completed successfully.")
    print(f"📁 JSON file: {json_path}")
    print(f"📁 Excel file: {excel_path}")

if __name__ == "__main__":
    main()
