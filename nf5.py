#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Census Extractor (nf5.py)
-------------------------
Reads Excel census sheets, splits them into manageable chunks,
sends them to the Groq LLM for JSON extraction, and outputs both
a JSON file and on-screen counts for employees vs dependents.

Requirements:
    pip install pandas openpyxl groq python-dotenv
"""

import os
import json
import time
import pandas as pd
from pathlib import Path
from groq import Groq
from openpyxl import load_workbook
from dotenv import load_dotenv
import warnings

warnings.filterwarnings("ignore")

# ----------------------------
# Load environment variables
# ----------------------------
load_dotenv()  # loads .env file automatically
api_key = os.getenv("GROQ_API_KEY")
if not api_key:
    raise ValueError("❌ GROQ_API_KEY not set. Ensure .env file exists and has GROQ_API_KEY=your_key")

# ----------------------------
# Configuration
# ----------------------------
MODEL_NAME = "llama-3.3-70b-versatile"
#CHUNK_SIZE = 5000
CHUNK_SIZE = 8000  # characters per chunk
OUTPUT_SUFFIX = "_employees.json"

client = Groq(api_key=api_key)

# ----------------------------
# Utility Functions
# ----------------------------
def read_excel_text(file_path):
    """Reads all sheets and concatenates them into plain text."""
    wb = load_workbook(filename=file_path, data_only=True)
    sheet_texts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            line = " | ".join([str(cell) if cell is not None else "" for cell in row])
            rows.append(line)
        text = "\n".join(rows)
        sheet_texts.append((sheet_name, text))
    return sheet_texts


def chunk_text(text, size=CHUNK_SIZE):
    """Splits text into smaller pieces."""
    return [text[i:i + size] for i in range(0, len(text), size)]


def clean_json_output(raw_text):
    """Attempts to extract valid JSON list from raw model output."""
    raw_text = raw_text.strip()
    if not raw_text:
        return []

    # Remove markdown fences if any
    if "```" in raw_text:
        parts = raw_text.split("```")
        if len(parts) >= 2:
            raw_text = parts[1]
        if raw_text.strip().startswith("json"):
            raw_text = raw_text.split("\n", 1)[1]

    raw_text = raw_text.strip()

    # Try parsing JSON directly
    try:
        data = json.loads(raw_text)
    except Exception:
        fixed = raw_text.replace("'", '"').replace("\n", " ")
        try:
            data = json.loads(fixed)
        except Exception:
            print("⚠️ Still invalid JSON, skipping this chunk.")
            return []

    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        return []

    return data


def extract_from_chunk(chunk_text, idx, total):
    """Send one chunk to Groq and return parsed data."""
    print(f"🧠 Sending LLM request part {idx}/{total}...")

    prompt = f"""
Extract employee census data and return ONLY valid JSON in the format:

[
  {{
    "first_name": "",
    "last_name": "",
    "relationship": "",
    "dependent": "",
    "medical_plan": "",
    "dental_plan": "",
    "vision_plan": ""
  }}
]

Input data:
{chunk_text}
"""

    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are a deterministic JSON-only extractor. "
                        "Return a valid JSON array only — no markdown or explanation. "
                        "If nothing found, return []."
                    )
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0,
            max_tokens=4096,
        )

        response_text = response.choices[0].message.content.strip()
        print(f"🧩 Raw output (first 300 chars):\n{response_text[:300]}")
        data = clean_json_output(response_text)
        return data

    except Exception as e:
        print(f"❌ Error during LLM request part {idx}: {e}")
        return []


# ----------------------------
# Main Execution
# ----------------------------
def main():
    import tkinter as tk
    from tkinter import filedialog
    tk.Tk().withdraw()

    file_path = filedialog.askopenfilename(
        title="Select Census Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    if not file_path:
        print("❌ No file selected.")
        return

    print(f"📘 Loaded workbook '{file_path}'")
    sheets = read_excel_text(file_path)
    all_records = []

    for sheet_name, text in sheets:
        chunks = chunk_text(text)
        print(f"📄 Sheet '{sheet_name}' -> {len(chunks)} chunks")

        for i, chunk in enumerate(chunks, start=1):
            data = extract_from_chunk(chunk, i, len(chunks))
            all_records.extend(data)

    # ----------------------------
    # Count and Output Results
    # ----------------------------
    if not all_records:
        print("❌ JSON file must contain a non-empty list.")
    else:
        employees = 0
        dependents = 0
        for rec in all_records:
            dep_flag = str(rec.get("dependent", "")).strip().lower()
            if dep_flag in ["y", "yes", "true", "1"]:
                dependents += 1
            else:
                employees += 1

        print(f"👩‍💼 Employees: {employees}")
        print(f"👨‍👩‍👧 Dependents: {dependents}")
        print(f"🎯 Total extracted records: {len(all_records)}")

    # Save output JSON
    out_path = Path(file_path).with_name(
        f"{Path(file_path).stem}{OUTPUT_SUFFIX}"
    )
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, indent=2, ensure_ascii=False)

    print(f"✅ Extraction completed.\nJSON file saved to: {out_path}")


if __name__ == "__main__":
    main()
